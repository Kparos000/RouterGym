"""Create blinded manual audit files from gold-matched production outputs."""

from __future__ import annotations

import argparse
import json
import random
from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd


REPO_ROOT = Path(__file__).resolve().parents[2]
GOLD_EVAL_DIR = REPO_ROOT / "RouterGym" / "results" / "analysis_outputs" / "gold_resolution_eval"
MATCHED_PATH = GOLD_EVAL_DIR / "gold_matched_production_outputs.jsonl"
OUTPUT_DIR = REPO_ROOT / "RouterGym" / "results" / "analysis_outputs" / "manual_audit"
DEFAULT_PER_CONFIG = 30
RANDOM_SEED = 42
COMPONENT_SCORE_COLUMNS = [
    "category_understanding_manual",
    "answer_actionable_manual",
    "answer_complete_manual",
    "resolution_steps_correct_manual",
    "escalation_appropriate_manual",
    "policy_grounded_manual",
]
OVERALL_SCORE_COLUMN = "overall_manual_quality"


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"Missing gold-matched outputs: {path}")
    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                parsed = json.loads(line)
                if isinstance(parsed, dict):
                    records.append(parsed)
    return records


def parse_config(config: str) -> dict[str, str]:
    tokens = str(config or "").split("__")
    parsed = {
        "router_family": tokens[0] if tokens else "",
        "base_model": "",
        "escalation_model": "",
        "memory_mode": "",
    }
    for token in tokens:
        if token.startswith("base_"):
            parsed["base_model"] = token.removeprefix("base_")
        elif token.startswith("esc_"):
            parsed["escalation_model"] = token.removeprefix("esc_")
        elif token.startswith("mem_"):
            parsed["memory_mode"] = token.removeprefix("mem_")
    return parsed


def is_bad(record: dict[str, Any]) -> bool:
    return (
        bool(record.get("parse_error"))
        or bool(record.get("validation_error"))
        or not bool(record.get("generation_valid", True))
    )


def bool_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes"}
    return bool(value)


def stringify(value: Any) -> str:
    if isinstance(value, str):
        return value
    if value is None:
        return ""
    return json.dumps(value, ensure_ascii=False)


def stable_all_gold_matched(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        records,
        key=lambda row: (
            str(row.get("analysis_config", "")),
            int(row.get("ticket_id", -1)) if str(row.get("ticket_id", "")).isdigit() else -1,
        ),
    )


def pick_stratified(records: list[dict[str, Any]], per_config: int) -> list[dict[str, Any]]:
    rng = random.Random(RANDOM_SEED)
    by_config: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        by_config[str(record.get("analysis_config", ""))].append(record)

    selected: list[dict[str, Any]] = []
    for config, config_records in sorted(by_config.items()):
        target = min(per_config, len(config_records))
        chosen: list[dict[str, Any]] = []

        bad_rows = [row for row in config_records if is_bad(row)]
        escalated_rows = [row for row in config_records if bool_value(row.get("escalated"))]
        non_escalated_rows = [row for row in config_records if not bool_value(row.get("escalated"))]

        for pool in (bad_rows, escalated_rows, non_escalated_rows):
            rng.shuffle(pool)
            for row in pool[: max(1, min(5, target // 6 or 1))]:
                if row not in chosen:
                    chosen.append(row)
                if len(chosen) >= target:
                    break
            if len(chosen) >= target:
                break

        by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in config_records:
            by_category[str(row.get("gold_topic_group", row.get("gold_label", "")))].append(row)
        while len(chosen) < target:
            added = False
            for category in sorted(by_category):
                pool = by_category[category]
                rng.shuffle(pool)
                for row in pool:
                    if row not in chosen:
                        chosen.append(row)
                        added = True
                        break
                if len(chosen) >= target:
                    break
            if not added:
                break

        selected.extend(chosen[:target])
    rng.shuffle(selected)
    return selected


def get_latency_ms(record: dict[str, Any]) -> Any:
    if record.get("latency_ms") is not None:
        return record.get("latency_ms")
    metrics = record.get("metrics")
    if isinstance(metrics, dict):
        return metrics.get("latency_ms")
    return None


def build_blinded_rows(selected: list[dict[str, Any]]) -> tuple[pd.DataFrame, pd.DataFrame]:
    configs = sorted({str(row.get("analysis_config", "")) for row in selected})
    rng = random.Random(RANDOM_SEED)
    anonymous_ids = [f"System {chr(65 + idx)}" for idx in range(len(configs))]
    rng.shuffle(anonymous_ids)
    config_to_anon = dict(zip(configs, anonymous_ids))

    blinded_rows: list[dict[str, Any]] = []
    key_rows: list[dict[str, Any]] = []
    for index, record in enumerate(selected, start=1):
        config = str(record.get("analysis_config", ""))
        gold_resolution = record.get("gold_resolution", {})
        if not isinstance(gold_resolution, dict):
            gold_resolution = {}
        audit_id = f"AUDIT-{index:04d}"
        blinded_rows.append(
            {
                "audit_id": audit_id,
                "anonymous_system_id": config_to_anon[config],
                "ticket_id": record.get("ticket_id", ""),
                "ticket_text": record.get("gold_ticket_text")
                or record.get("original_query")
                or record.get("ticket_request", ""),
                "gold_label": record.get("gold_topic_group") or record.get("gold_label", ""),
                "gold_resolution_summary": gold_resolution.get("summary", ""),
                "gold_resolution_steps": stringify(gold_resolution.get("steps", [])),
                "gold_acceptance_criteria": stringify(
                    gold_resolution.get("acceptance_criteria", [])
                ),
                "generated_final_answer": record.get("final_answer", ""),
                "generated_resolution_steps": stringify(record.get("resolution_steps", [])),
                "generated_reasoning": record.get("reasoning", ""),
                "generated_escalation_flags": stringify(record.get("escalation_flags", {})),
                **{column: "" for column in COMPONENT_SCORE_COLUMNS},
                OVERALL_SCORE_COLUMN: "",
                "reviewer_id": "",
                "reviewer_notes": "",
            }
        )
        parsed = parse_config(config)
        key_rows.append(
            {
                "audit_id": audit_id,
                "anonymous_system_id": config_to_anon[config],
                "analysis_config": config,
                "router_family": record.get("router_family") or parsed["router_family"],
                "base_model": record.get("base_model") or parsed["base_model"],
                "escalation_model": record.get("escalation_model") or parsed["escalation_model"],
                "memory_mode": record.get("memory_mode") or parsed["memory_mode"],
                "ticket_id": record.get("ticket_id", ""),
                "gold_label": record.get("gold_topic_group") or record.get("gold_label", ""),
                "parse_error_present": bool(record.get("parse_error")),
                "validation_error_present": bool(record.get("validation_error")),
                "generation_valid": record.get("generation_valid"),
                "escalated": record.get("escalated"),
                "total_cost_usd": record.get("total_cost_usd"),
                "latency_ms": get_latency_ms(record),
            }
        )
    return pd.DataFrame(blinded_rows), pd.DataFrame(key_rows)


def write_excel(blinded: pd.DataFrame, rubric_text: str, path: Path) -> bool:
    try:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return False

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        blinded.to_excel(writer, index=False, sheet_name="Review")
        rubric_lines = [{"Rubric": line} for line in rubric_text.splitlines()]
        pd.DataFrame(rubric_lines).to_excel(writer, index=False, sheet_name="Rubric")

    workbook = load_workbook(path)
    review_sheet = workbook["Review"]
    review_sheet.freeze_panes = "D2"
    review_sheet.auto_filter.ref = review_sheet.dimensions

    content_fill = PatternFill("solid", fgColor="EAF2F8")
    score_fill = PatternFill("solid", fgColor="FFF2CC")
    notes_fill = PatternFill("solid", fgColor="E2F0D9")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    score_header_fill = PatternFill("solid", fgColor="FFE699")

    widths = {
        "audit_id": 14,
        "anonymous_system_id": 18,
        "ticket_id": 12,
        "ticket_text": 50,
        "gold_label": 22,
        "gold_resolution_summary": 50,
        "gold_resolution_steps": 55,
        "gold_acceptance_criteria": 55,
        "generated_final_answer": 60,
        "generated_resolution_steps": 55,
        "generated_reasoning": 45,
        "generated_escalation_flags": 35,
        "reviewer_id": 16,
        "reviewer_notes": 45,
    }
    score_widths = {column: 20 for column in COMPONENT_SCORE_COLUMNS}
    score_widths[OVERALL_SCORE_COLUMN] = 22
    score_widths["reviewer_id"] = 18
    widths.update(score_widths)

    header_to_col = {cell.value: cell.column_letter for cell in review_sheet[1]}
    for header, width in widths.items():
        column_letter = header_to_col.get(header)
        if column_letter:
            review_sheet.column_dimensions[column_letter].width = width

    for row in review_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in review_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = (
            score_header_fill
            if cell.value in {*COMPONENT_SCORE_COLUMNS, OVERALL_SCORE_COLUMN}
            else header_fill
        )

    content_columns = [
        "ticket_text",
        "gold_resolution_summary",
        "gold_resolution_steps",
        "gold_acceptance_criteria",
        "generated_final_answer",
        "generated_resolution_steps",
        "generated_reasoning",
        "generated_escalation_flags",
    ]
    for header in content_columns:
        column_letter = header_to_col.get(header)
        if column_letter:
            for cell in review_sheet[f"{column_letter}"][1:]:
                cell.fill = content_fill
    for header in [*COMPONENT_SCORE_COLUMNS, OVERALL_SCORE_COLUMN]:
        column_letter = header_to_col.get(header)
        if column_letter:
            for cell in review_sheet[f"{column_letter}"][1:]:
                cell.fill = score_fill
    for header in ["reviewer_id", "reviewer_notes"]:
        column_letter = header_to_col.get(header)
        if column_letter:
            for cell in review_sheet[f"{column_letter}"][1:]:
                cell.fill = notes_fill

    for row_idx in range(2, review_sheet.max_row + 1):
        review_sheet.row_dimensions[row_idx].height = 92

    last_row = review_sheet.max_row
    component_validation = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    component_validation.error = "Use 0, 1, or 2 for component scores."
    component_validation.prompt = "Select 0, 1, or 2."
    component_validation.showDropDown = False
    review_sheet.add_data_validation(component_validation)
    for column in COMPONENT_SCORE_COLUMNS:
        column_letter = header_to_col.get(column)
        if column_letter:
            component_validation.add(f"{column_letter}2:{column_letter}{last_row}")

    overall_validation = DataValidation(
        type="list",
        formula1='"0,1,2,3,4,5,6,7,8,9,10"',
        allow_blank=True,
    )
    overall_validation.error = "Use a whole number from 0 to 10."
    overall_validation.prompt = "Select an overall score from 0 to 10."
    overall_validation.showDropDown = False
    review_sheet.add_data_validation(overall_validation)
    overall_col = header_to_col.get(OVERALL_SCORE_COLUMN)
    if overall_col:
        overall_validation.add(f"{overall_col}2:{overall_col}{last_row}")

    reviewer_validation = DataValidation(
        type="list",
        formula1='"Ogaga,Reviewer_2,Reviewer_3,Reviewer_4"',
        allow_blank=True,
    )
    reviewer_validation.error = "Select one of the reviewer IDs or leave blank until assigned."
    reviewer_validation.prompt = "Select reviewer ID."
    reviewer_validation.showDropDown = False
    review_sheet.add_data_validation(reviewer_validation)
    reviewer_col = header_to_col.get("reviewer_id")
    if reviewer_col:
        reviewer_validation.add(f"{reviewer_col}2:{reviewer_col}{last_row}")

    comments = {
        "category_understanding_manual": "0=wrong/missing, 1=partial, 2=good category and user-need understanding.",
        "answer_actionable_manual": "0=not actionable, 1=partially actionable, 2=clear operational actions.",
        "answer_complete_manual": "0=major omissions, 1=some omissions, 2=covers key gold requirements.",
        "resolution_steps_correct_manual": "0=wrong/missing steps, 1=partly correct, 2=correct relevant steps.",
        "escalation_appropriate_manual": "0=wrong escalation decision, 1=unclear/partial, 2=appropriate escalation handling.",
        "policy_grounded_manual": "0=unsupported or contradicts policy, 1=partly grounded, 2=policy/KB aligned.",
        "overall_manual_quality": "Overall 0-10 quality: 0 useless, 5 partially useful, 10 excellent/correct/actionable/complete.",
        "reviewer_id": "Select your reviewer ID.",
        "reviewer_notes": "Optional free-text notes. No dropdown is applied.",
    }
    for header, text in comments.items():
        column_letter = header_to_col.get(header)
        if column_letter:
            review_sheet[f"{column_letter}1"].comment = Comment(text, "RouterGym")

    table_ref = review_sheet.dimensions
    table = Table(displayName="ManualAuditReview", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    review_sheet.add_table(table)

    rubric_sheet = workbook["Rubric"]
    rubric_sheet.column_dimensions["A"].width = 120
    for row in rubric_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    rubric_sheet["A1"].font = Font(bold=True)

    workbook.save(path)
    return True


def write_rubric(path: Path) -> None:
    content = """# Manual Audit Rubric

Use this rubric to score generated resolution quality against the gold reference. Do not open or
inspect `manual_audit_key.csv` or `manual_audit_full_key.csv` until after scoring is complete.

## Component Score Columns

- `category_understanding_manual`: Does the generated response understand the ticket category and user need?
- `answer_actionable_manual`: Does it give concrete actions that could be followed by support staff or the user?
- `answer_complete_manual`: Does it cover the important parts of the gold resolution and avoid major omissions?
- `resolution_steps_correct_manual`: Are the generated resolution steps correct, ordered, and relevant?
- `escalation_appropriate_manual`: Does it escalate when escalation is needed and avoid unnecessary escalation?
- `policy_grounded_manual`: Is it consistent with the supplied policies, KB expectations, and acceptance criteria?

Score each component as:

- 0 = poor, wrong, unsafe, missing, or not usable
- 1 = partial or acceptable but incomplete
- 2 = good, correct, useful, and operationally actionable

## Overall Score

`overall_manual_quality` should be scored from 0 to 10:

- 0 = completely wrong, unsafe, or useless
- 1-2 = incorrect or mostly unusable resolution
- 3-4 = weak resolution with major omissions
- 5 = partially useful but incomplete
- 6 = useful but clearly missing important detail
- 7-8 = good resolution suitable for most operational use
- 9 = excellent resolution with only minor differences from the gold reference
- 10 = excellent, correct, actionable, complete, and closely aligned with the gold reference

## Consistency Guidance

Reviewers should compare the generated answer and generated resolution steps against the gold
summary, gold steps, and acceptance criteria. Penalize hallucinated steps, missing required actions,
incorrect escalation decisions, and vague answers that are not operationally useful.

The Excel workbook is the primary review artifact. Use the dropdowns in the `Review` sheet for all
component scores, `overall_manual_quality`, and `reviewer_id`. Do not manually add scoring columns
or decode system identities while reviewing.

The anonymous system ID must not be decoded during review. If multiple reviewers score the file,
each reviewer should fill `reviewer_id` so inter-reviewer agreement can be computed later.
"""
    path.write_text(content, encoding="utf-8")


def write_csv_or_warn(df: pd.DataFrame, path: Path, label: str) -> bool:
    try:
        df.to_csv(path, index=False)
    except PermissionError:
        print(f"{label} CSV locked and was not overwritten: {path}")
        print("Close the file in Excel and rerun this command to refresh the CSV.")
        return False
    return True


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--all-gold-matched",
        action="store_true",
        help="Create a blinded audit file for every gold-matched production output.",
    )
    parser.add_argument(
        "--per-config",
        type=int,
        default=DEFAULT_PER_CONFIG,
        help=f"Rows per config for the default stratified sample. Default: {DEFAULT_PER_CONFIG}.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    records = read_jsonl(MATCHED_PATH)
    selected = (
        stable_all_gold_matched(records)
        if args.all_gold_matched
        else pick_stratified(records, args.per_config)
    )
    blinded, key = build_blinded_rows(selected)
    if args.all_gold_matched:
        blinded_path = OUTPUT_DIR / "manual_audit_full_blinded.csv"
        key_path = OUTPUT_DIR / "manual_audit_full_key.csv"
        xlsx_path = OUTPUT_DIR / "manual_audit_full_blinded.xlsx"
    else:
        blinded_path = OUTPUT_DIR / "manual_audit_blinded.csv"
        key_path = OUTPUT_DIR / "manual_audit_key.csv"
        xlsx_path = None

    blinded_csv_written = write_csv_or_warn(blinded, blinded_path, "Blinded audit")
    key_csv_written = write_csv_or_warn(key, key_path, "Manual audit key")
    rubric_path = OUTPUT_DIR / "manual_audit_rubric.md"
    write_rubric(rubric_path)
    rubric_text = rubric_path.read_text(encoding="utf-8")

    xlsx_created = False
    if xlsx_path is not None:
        xlsx_created = write_excel(blinded, rubric_text, xlsx_path)

    print("Manual audit sample created")
    print(f"Rows: {len(blinded)}")
    print(f"Configs: {key['analysis_config'].nunique()}")
    print("Rows per anonymous system:")
    print(blinded["anonymous_system_id"].value_counts().sort_index().to_string())
    print(f"Blinded CSV: {blinded_path} ({'written' if blinded_csv_written else 'locked'})")
    print(f"Key CSV: {key_path} ({'written' if key_csv_written else 'locked'})")
    if xlsx_path is not None:
        if xlsx_created:
            print(f"Excel workbook: {xlsx_path}")
        else:
            print("Excel workbook skipped: install openpyxl to enable .xlsx output.")
    print(f"Outputs written to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
